'''
Brianna Singer
July 5th, 2017
Version: 1.0

'''

#to call from terminal:
# first go into the directory that holds this file, probably "Scraper"
# type: $ python -c 'import netGetter; print netGetter.getNets("John", "Stewart")'
# or just type: python netGetter.py to activate the custom commands


import requests
from bs4 import BeautifulSoup
import openpyxl


'''takes input of first name and last name and returns a dictionary of result names and NetIDs'''
def getNets(first, last):

	url = 'https://www.cornell.edu/search/people.cfm?q=' + str(first) + '+' + str(last) + '&tab=people'
	r = requests.get(url)
	soup = BeautifulSoup(r.content)
	if not soup.find_all('a', {'name': 'Alumni'}): #in case there is no alumni by that name on the website
		return "No such alumni in database, yo"
	alumLink = soup.find_all('a', {'name': 'Alumni'})[0]
	alumCap = alumLink.parent
	resultTable = alumCap.parent
	if resultTable['class'] == 'results cu-table':
		rows = resultTable.find_all("tr")
	else:
		print "error finding result table"

	resultDict = dict()	

	for row in rows[1:]:
		rowLinks = row.find_all('a') 
		nameLink = rowLinks[0]
		contactName = nameLink.text
		#print contactName
		netTD = row.find_all('td')[2] #ex: <td><span class="label">NetID:</span> jbs247</td>
		netRaw = netTD.text
		if 'NetID: ' in netRaw:
			contactNet = netRaw.replace('NetID: ', '')
			#print contactNet
			resultDict[contactName] = contactNet

		else:
			print 'could not find netID td tag, yo'

	return resultDict




''' Takes a first name, a middle name, a last name, and a dictionary of names and netIDs as parameters. 
Returns a new version of the dictionary with only names that match the first and last 
name parameters. Allows middle names and middle initials.'''

def netNarrow(first, last, data): #takes a dictionary as a parameter

	# firstName = 'James'
	# lastName = 'Smith'

	firstName = first
	lastName = last
	netData = data



	removables = []

	removables.append(' PhD')
	removables.append(' Esq')
	removables.append(' Jr')
	removables.append(' III')
	removables.append(' II')
	removables.append(' I')
	removables.append(',')

	matchDict = {}


	for key in netData: # key represents just the name, not the netID
		newKey = key
		for removable in removables:
			if removable in newKey:
				newKey = newKey.replace(removable, '') #take the removable out of key
		#print newKey

		frontLen = len(firstName)
		backLen = len(lastName)

		frontSec = newKey[:frontLen] # this is the first several letters of key
		backSec = newKey[-backLen:] # this is the last several letters of key

		if frontSec == firstName and backSec == lastName:
			print key + 'has first name: ' + firstName + 'and last name ' + lastName
			matchDict[key] = netData[key]
		else:
			print "not a match"

	return matchDict


	'''The parameters consist of one contact persons contactID, a dictionary of names and NetIDs
	for this one contact person as "data," and a file "file". For each item in data,
	it writes a row into "nfile." ''' 
def storeNets(cID, data, cfile, nfile):

	contactID = cID
	contactDict = data
	netFileString = nfile
	conFileString = cfile

	netFileReadable = openpyxl.load_workbook(netFileString, read_only = True)
	netFileWritable = openpyxl.load_workbook(netFileString, data_only = True)
	rSheet = netFileReadable.get_sheet_by_name('nSheet')
	wSheet = netFileWritable.get_sheet_by_name('nSheet')

	# create a dictionary for nfile: Key: column name, value: column letter
	#the following are columns in 'nets.xlsx' unless it has been changed
	#contactID	netID	cornellName	linkFirstName	linkLastName	firm

	colNames = dict()
	alphabetList = ['A', 'B', 'C', 'D', 'E', 'F', 
	'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
	'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V',
	'W', 'X', 'Y', 'Z']

	for i in range(0, len(rSheet[1])):
		colName =  rSheet[1][i].value
		colNames[colName] = alphabetList[i]

	# here is what colNames should look like at this point:
	#{u'firm': 'F', u'netID': 'B', u'linkFirstName': 'D', 
	# u'cornellName': 'C', u'contactID': 'A', u'linkLastName': 'E'}

	# loop through contactDic, for every entry, make a new row in netFile as follows:

	# contactID: contactID,	netID: dict value,	
	# cornellName: dict key, linkFirstName: use findContactInfo(), 
	# linkLastName: use findContactInfo(), firm: use findContactInfo()

	for k, v in contactDict.items():
		#print "key is " + k + " and value is " + v

		# use this goldRow for each cell value:
		goldRow = wSheet.max_row + 1 #this should be the next blank row in the spreadsheet

		#write contactID value
		cIDValue = contactID
		cIDLetter = colNames['contactID']
		wSheet[cIDLetter + str(goldRow)] = cIDValue


		#write contactID value
		netIDValue = v #v is the variable representing the current value in the dictionary loop
		netIDLetter = colNames['netID']
		wSheet[netIDLetter + str(goldRow)] = netIDValue


		#write cornellName value
		cornellNameValue = k #k is the variable representing the current key in the dictionary loop
		cornellNameLetter = colNames['cornellName']
		wSheet[cornellNameLetter + str(goldRow)] = cornellNameValue

		#write linkFirstName value
		linkFirstNameValue = findContactInfo(conFileString, contactID, 'firstName')
		linkFirstNameLetter = colNames['linkFirstName']
		wSheet[linkFirstNameLetter + str(goldRow)] = linkFirstNameValue

		#write linkLastName value
		linkLastNameValue = findContactInfo(conFileString, contactID, 'lastName')
		linkLastNameLetter = colNames['linkLastName']
		wSheet[linkLastNameLetter + str(goldRow)] = linkLastNameValue

		#write firm value
		firmValue = findContactInfo(conFileString, contactID, 'firm')
		firmLetter = colNames['firm']
		wSheet[firmLetter + str(goldRow)] = firmValue

	# save and close netFile

	netFileWritable.save(conFileString)

	return


	''' This function opens a csv file which has two rows: contactID and NetID. It returns all of the
	NetIDs that match the given cID. '''

def getNetIDs(cID):
	return "Error, yo: function not defined yet"

	''' Goes into a comma separated file "file." Finds the colum with the title "cat" and returns 
	its value on the row that also includes cID'''

def findContactInfo(file, cID, cat): # string name of .xlsx file, integer of contactID, string of column name

	contactsFile = file # 'contacts.xlsx', most likely
	contactID = cID
	columnCat = cat

	cFile = openpyxl.load_workbook(contactsFile, read_only= True) # just got this line working,
	sheet = cFile.get_sheet_by_name('cSheet')

	# make a dictionary of all the column names and their respective letters

	colNames = dict()
	alphabetList = ['A', 'B', 'C', 'D', 'E', 'F', 
	'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
	'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V',
	'W', 'X', 'Y', 'Z']

	for i in range(0, len(sheet[1])):
		colName =  sheet[1][i].value
		colNames[colName] = alphabetList[i]

	#Here is what colNames should look like now:
	 #{u'firm': 'B',
	 # u'firstName': 'D',
	 # u'title': 'G',
	 # u'gender': 'C',
	 # u'contactID': 'A',
	 # u'gradYear': 'H',
	 # u'lastName': 'F',
	 # u'midName': 'E'}

	# search the dictionary for a key that has the same name as glob-variable "cat"
	# if there is no such key, return an error message

	try: 
		colLet = colNames[columnCat] #the single '=' here is on purpose
	except:
		return "Error, yo: 'cat' parameter does not match any columns in the 'file' spreadsheet"

	# find the appropriate row with the specified cID

	contactIDColLet = colNames['contactID'] #this should be 'A' unless the spreadsheet is changed
	goldRow = contactID + 1 # This should be the row that contains the contactID

	if sheet[contactIDColLet + str(goldRow)].value == contactID:
		# returns the value at specified row for specified category
		return sheet[colLet + str(goldRow)].value 
	else:
		return "Error, yo: Couldn't find that contactID in given 'file'"




	'''Takes a given cID and goes into a comma separated file "file" to find the row with that 
	contactID. It extracts the first and last name from that row and uses getNets() and netNarrow() 
	to produce a dictionary of netIDs. It returns this Dictionary. '''

def getNetsByCID(cID, file):
	return "Error, yo: function not defined yet"

	'''This funciton helps sendEmail() keep a log of messages sent by storing records in a csv file.'''

def logMessage(date, recName, recAddress):
	return "Error, yo: function not defined yet"

	'''This function takes inputs for recipient email address, gender last name, and firm. It then 
	sends a customized email message from the sender address bls252@cornell.edu. It also uses a 
	helper function to log the email in a csv file. '''

def sendEmail(recAddress, gender, lName, firm):
	return "Error, yo: function not defined yet"

# ///////////////////////////// Custom Commands ///////////////////////////////

# search1 = getNets("Dennis", "Kim")

# narrowResults = netNarrow("Dennis", "Kim", search1)

# print narrowResults


# coolLastName = findContactInfo('contacts.xlsx', 38, 'firm')

# print coolLastName # and what a cool last name it is :)

tcontactID = 38 #cID
tcontactDict = {u'Dennis Yun Kim': u'dyk22', u'Dennis Sekwon Kim': u'dsk35'} #data
tnetFileString = 'nets.xlsx' #nfile
tconFileString = 'contacts.xlsx' #cfile

storeNets(tcontactID, tcontactDict, tconFileString, tnetFileString)


