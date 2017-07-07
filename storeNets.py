# make sure to enter the directory "scraper" first

import openpyxl

#the following are columns in 'nets.xlsx'
#contactID	netID	cornellName	linkName	firm


	'''The parameters consist of one contact persons contactID, a dictionary of names and NetIDs
	for this one contact person as "data," and a file "file". For each item in data,
	it writes a row into "nfile." ''' 
def storeNets(cID, data, cfile, nfile):

	contactID = 38 #cID
	contactDict = {u'Dennis Yun Kim': u'dyk22', u'Dennis Sekwon Kim': u'dsk35'} #data
	netFileString = 'nets.xlsx' #nfile
	conFileString = 'contacts.xlsx' #cfile

	netFileReadable = openpyxl.load_workbook(netFileString, read_only = True)
	netFileWritable = openpyxl.load_workbook(netFileString, data_only = True)
	rSheet = netFileReadable.get_sheet_by_name('nSheet')
	wSheet = netFileWritable.get_sheet_by_name('nSheet')

	# create a dictionary for nfile: Key: column name, value: column letter
	#the following are columns in 'nets.xlsx' unless it has been changed
	#contactID	netID	cornellName	linkFirstName	linkLastName	firm

	colNames = dict()
	alphabetList = ['A', 'B', 'C', 'D', 'E', 'F', 
	'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
	'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V',
	'W', 'X', 'Y', 'Z']

	for i in range(0, len(rSheet[1])):
		colName =  rSheet[1][i].value
		colNames[colName] = alphabetList[i]

	# here is what colNames should look like at this point:
	#{u'firm': 'F', u'netID': 'B', u'linkFirstName': 'D', 
	# u'cornellName': 'C', u'contactID': 'A', u'linkLastName': 'E'}

	# loop through contactDic, for every entry, make a new row in netFile as follows:

	# contactID: contactID,	netID: dict value,	
	# cornellName: dict key, linkFirstName: use findContactInfo(), 
	# linkLastName: use findContactInfo(), firm: use findContactInfo()

	# for k, v in contactDict.items():
	# 	print "key is " + k + " and value is " + v


	# use this goldRow for each cell value:
		goldRow = wSheet.max_row + 1 #this should be the next blank row in the spreadsheet

		#write contactID value
		cIDValue = contactID
		cIDLetter = colNames['contactID']
		wSheet[cIDLetter + str(goldRow)] = cIDValue


		#write contactID value
		netIDValue = v #v is the variable representing the current value in the dictionary loop
		netIDLetter = colNames['netID']
		wSheet[netIDLetter + str(goldRow)] = netIDValue


		#write cornellName value
		cornellNameValue = k #k is the variable representing the current key in the dictionary loop
		cornellNameLetter = colNames['cornellName']
		wSheet[cornellNameLetter + str(goldRow)] = cornellNameValue

		#write linkFirstName value
		linkFirstNameValue = findContactInfo(conFileString, contactID, 'firstName')
		linkFirstNameLetter = colNames['linkFirstName']
		wSheet[linkFirstNameLetter + str(goldRow)] = linkFirstNameValue

		#write linkLastName value
		linkLastNameValue = findContactInfo(conFileString, contactID, 'lastName')
		linkLastNameLetter = colNames['linkLastName']
		wSheet[linkLastNameLetter + str(goldRow)] = linkLastNameValue

		#write firm value
		firmValue = findContactInfo(conFileString, contactID, 'firm')
		firmLetter = colNames['firm']
		wSheet[firmLetter + str(goldRow)] = firmValue


# save and close netFile

netFileWritable.save(conFileString)

# /////////////////////////Testing Only /////////////////////////////////////

for row in sheet.rows:
    for cell in row:
        print(cell.value)

for row in sheet.rows:
    print row[0].value

print sheet['D1'].value


#Colums, their letters, and the kind of info they contain
contactIDCol = 'A'
firmCol = 'B'
genderCol = 'C'
firstNameCol = 'D'
middleCol = 'E'
lastNameCol = 'F'
jobTitleCol = 'G'
gradYearCol = 'H'

print sheet[goldRow][0].value # 1st row, A column



print sheet[Dvar + str(goldRow)].value


cFile = openpyxl.load_workbook('contacts.xlsx', data_only= True) 

sheet = cFile.get_sheet_by_name('cSheet')
sheet['I1'] = 'hello world'
cFile.save('contacts.xlsx')
